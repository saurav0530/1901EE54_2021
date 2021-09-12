import os
import csv
from openpyxl import load_workbook, Workbook
os.system("cls")
try:
    os.mkdir("./output_by_subject")
except FileExistsError:
    pass
try:
    os.mkdir("./output_individual_roll")
except FileExistsError:
    pass

def addData(folder,filename,data):
    try:
        wb = load_workbook('./'+folder+'/'+filename+'.xlsx')
        ws = wb.active
        ws.append(data)
        wb.save('./'+folder+'/'+filename+'.xlsx')
        wb.close()
    except FileNotFoundError:
        wb=Workbook()
        ws=wb.active
        ws.title=filename
        headers = ["rollno","register_sem","subno","sub_type"]
        ws.append(headers)
        ws.append(data)
        wb.save('./'+folder+'/'+filename+'.xlsx')
        wb.close()

# Function which extracts data from given csv file and add it to corresponding file of corresponding folder
def output_by_subject():
    with open('regtable_old.csv', 'r') as csvfile:
        reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)  # Both
        for data in reader:
            results = [data[0],data[1],data[3],data[8]]
            if(results[0]=="rollno"):
                continue
            addData('output_by_subject',results[2],results)      # adds data of each file of output_by_subject

def output_individual_roll():
    with open('regtable_old.csv', 'r') as csvfile:
        reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)  # Both
        for data in reader:
            results = [data[0],data[1],data[3],data[8]]
            if(results[0]=="rollno"):
                continue
            addData('output_individual_roll',results[0],results)       # adds data of each file of output_individual_roll

# Function call made
output_by_subject()
output_individual_roll()