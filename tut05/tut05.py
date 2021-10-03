import os
import csv
import shutil
from openpyxl import load_workbook, Workbook
os.system("cls")

# Checking/Creating output directory
try:
    os.mkdir("./output")
except FileExistsError:
    shutil.rmtree("./output")
    os.mkdir("./output")
    pass

# Grade to pointer relation dictionary declaration
grd_dict={
    'AA':10,'AA*':10,
    'AB':9,'AB*':9,
    'BB':8,'BB*':8,
    'BC':7,'BC*':7,
    'CC':6,'CC*':6,
    'CD':5,'CD*':5,
    'DD':4,'DD*':4,
    'F':0,'F*':0,
    'I':0,'I*':0
}
# Dictionary declaration to store subject data of subjects-master.csv
subj_dict = dict()

# Dictionary declaration to store student data of names-roll.csv
name_dict = dict()

# Function to fill subj_dict dictionary
with open('subjects_master.csv', 'r') as csvfile:
    reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
    for data in reader:
        subj_dict[data[0]]={
            'name' : data[1],
            'ltp' : str(data[2])
        }

# Function to fill name_dict dictionary
with open('names-roll.csv', 'r') as csvfile:
    reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
    for data in reader:
        name_dict[data[0]]=data[1]

# Function to fill student's semester data to corresponding workbook's sheet
def add_sem_to_sheets(sem_no, roll, semData):
    ws=5
    wb=5
    # list to store data to be appended to one row of a semester sheet
    dataToBeAdded = [1, semData['no'],subj_dict[semData['no']]['name'], subj_dict[semData['no']]['ltp'], semData['credit'], semData['type'], semData['grade']]
    
    # loading required workbook
    try:
        wb=load_workbook('./output/'+roll+'.xlsx')
    except FileNotFoundError:
        wb=Workbook()

    # loading required worksheet and appending rows to it
    try:
        ws=wb['Sem'+sem_no]
        dataToBeAdded[0]=ws.max_row
        ws.append(dataToBeAdded)
        wb.save('./output/'+roll+'.xlsx')
    except:
        ws=wb.create_sheet('Sem'+sem_no)
        ws.append(['Sl No.','Subject No.','Subject Name', 'L-T-P', 'Credit', 'Subject Type', 'Grade'])
        ws.append(dataToBeAdded)
        wb.save('./output/'+roll+'.xlsx')

# Function to fill Overall sheet of a student
def calculate_grade(filename):
    # loading workbook of corresponding student
    wb=load_workbook('./output/'+filename+'.xlsx')
    number_of_sheets = len(wb.sheetnames)
    
    # declaring list to store credits, cum_credits, spi and cpi
    semester_array = ['Semester No.']
    credits_array = [0]
    spi_array = ['SPI']
    cumulative_credit_array = [0]
    cpi_array = [0]

    # filling semester number lists data
    for ws in wb.worksheets:
        if(ws.title=='Sheet'):
            continue
        semester_array.append((int)(ws.title.split('m')[1]))

    # filling all lists data
    for i in semester_array:
        if i=='Semester No.':
            continue
        ws=wb['Sem'+str(i)]
        credits=0
        prod=1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            prod += ((int)(row[4].value))*grd_dict[row[6].value]
            credits += (int)(row[4].value)
        spi = prod
        credits_array.append(credits)
        spi_array.append(spi)
        cumulative_credit_array.append(cumulative_credit_array[-1]+credits)
    for i in range(1,number_of_sheets):
        temp=0
        for j in range(1,i+1):
            temp += (int)(spi_array[j])
        cpi_array.append(temp)
    for i in range(1,number_of_sheets):
        spi_array[i]= (float)("{0:.2f}".format(((spi_array[i]*100)/credits_array[i])*0.01))
        cpi_array[i] = (float)("{0:.2f}".format(((cpi_array[i]*100)/cumulative_credit_array[i])*0.01))
    
    # Modifying each lists base index with its description name
    credits_array[0]='Semester Wise Credit Taken'
    spi_array[0]='SPI'
    cpi_array[0]='CPI'
    cumulative_credit_array[0]='Total Credits Taken'
    ws=wb['Sheet']
    ws.title = 'Overall'

    # Adding all data to overall sheet
    ws.append(['Roll No.', filename])
    ws.append(['Name of Student', name_dict[filename]])
    ws.append(['Discipline', ((str)(filename))[4]+((str)(filename))[5]])
    ws.append(semester_array)
    ws.append(credits_array)
    ws.append(spi_array)
    ws.append(cumulative_credit_array)
    ws.append(cpi_array)

    # Saving each file
    wb.save('./output/'+filename+'.xlsx')

# Main function definition
def generate_marksheet():
    with open('./grades.csv', 'r') as csvfile:
        reader = csv.reader(csvfile, delimiter=',', skipinitialspace=True)
        for data in reader:
            if(data[0]=='Roll'):
                continue
            semData = {
                'no': data[2],
                'credit' : data[3],
                'type' : data[5],
                'grade' : data[4]
            }

            # Calling function to add semester data to each student's file
            add_sem_to_sheets(data[1],data[0],semData)
    
    for file in os.listdir('./output'):
        filename = file.split('.')[0]
        
        # Calling function to fill overall sheet
        calculate_grade(filename)

# Calling main function
generate_marksheet()